# -*- coding: utf-8 -*-
"""クロネコB2(ヤマト運輸の送り状取込システム)向け、DM便の書き出し。

店で使っているB2の取込テンプレート(xlsx)は列位置で読むので、列名より
「列の位置が同じか」が大事。2026-08-06、実際に使っているテンプレートを
見せてもらい、必須項目の列位置を確認して作った。
"""
import datetime
import io

import openpyxl

import store_info

# B2テンプレートの列位置(1-indexed)。ここだけがテンプレートと合っていればよい。
COL_KIND = 2        # 送り状種類(3=DM便)
COL_SHIP_DATE = 5   # 出荷予定日
COL_TEL = 9         # お届け先電話番号
COL_POSTAL = 11     # お届け先郵便番号
COL_ADDRESS = 12    # お届け先住所
COL_ADDRESS2 = 13   # お届け先アパートマンション名
COL_NAME = 16       # お届け先名
COL_HONORIFIC = 18  # 敬称
COL_SENDER_POSTAL = 22   # ご依頼主郵便番号
COL_SENDER_ADDRESS = 23  # ご依頼主住所
COL_SENDER_NAME = 25     # ご依頼主名

# 見出し行(空欄の列は元テンプレートに合わせて「空欄」と表示するだけ)
_HEADERS = {
    COL_KIND: "送り状種類(3=DM便)",
    COL_SHIP_DATE: "出荷予定日",
    COL_TEL: "お届け先電話番号",
    COL_POSTAL: "お届け先郵便番号",
    COL_ADDRESS: "お届け先住所",
    COL_ADDRESS2: "お届け先アパートマンション名",
    COL_NAME: "お届け先名",
    COL_HONORIFIC: "敬称",
    COL_SENDER_POSTAL: "ご依頼主郵便番号",
    COL_SENDER_ADDRESS: "ご依頼主住所",
    COL_SENDER_NAME: "ご依頼主名",
}
_LAST_COL = 25


def export_bytes(customers, ship_date=None):
    """customers = [{"name","tel","postal","address","address2"}, ...] からxlsxバイト列を作る。"""
    ship_date = ship_date or datetime.date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col in range(1, _LAST_COL + 1):
        ws.cell(row=1, column=col, value=_HEADERS.get(col, "空欄"))

    for i, c in enumerate(customers, start=2):
        ws.cell(row=i, column=COL_KIND, value=3)  # DM便固定
        date_cell = ws.cell(row=i, column=COL_SHIP_DATE, value=ship_date)
        date_cell.number_format = "yyyy/mm/dd"
        ws.cell(row=i, column=COL_TEL, value=c["tel"])
        ws.cell(row=i, column=COL_POSTAL, value=c["postal"])
        ws.cell(row=i, column=COL_ADDRESS, value=c["address"])
        if c.get("address2"):
            ws.cell(row=i, column=COL_ADDRESS2, value=c["address2"])
        ws.cell(row=i, column=COL_NAME, value=c["name"])
        ws.cell(row=i, column=COL_HONORIFIC, value="様")
        ws.cell(row=i, column=COL_SENDER_POSTAL, value=store_info.SENDER_POSTAL)
        ws.cell(row=i, column=COL_SENDER_ADDRESS, value=store_info.SENDER_ADDRESS)
        ws.cell(row=i, column=COL_SENDER_NAME, value=store_info.STORE_NAME)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
