# -*- coding: utf-8 -*-
"""11台帳(xlsx)の「構造」だけを調べる診断ツール(個人情報は表示しない)。

各ファイルのシート名・行数・列数と、1行目(見出し候補)のセル数だけを表示する。
中身の値は一切表示しないので、実データに対して実行しても安全。

★取込のあとの「答え合わせ」に使う。ここで出る台帳ごとの行数と、
import_csv.py が最後に出す取り込み件数を見比べれば、取りこぼしが分かる。
(販売台帳・処方箋は数え方が違うのでズレて正常。docs/migration-guide.md 参照)

使い方:
  python3 scripts/diag_real_xlsx.py            # data/real/xlsx(無ければ data/real 直下)
  python3 scripts/diag_real_xlsx.py --demo     # data/demo で動作確認
  python3 scripts/diag_real_xlsx.py <フォルダ>  # 任意フォルダ
"""
import glob
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _paths import xlsx_dir   # noqa: E402  置き場の決め方は _paths.py にまとめてある

try:
    import openpyxl
except ImportError:
    print("[エラー] openpyxl が必要です。 pip3 install openpyxl を実行してください。")
    sys.exit(1)

args = [a for a in sys.argv[1:] if a != "--demo"]
if "--demo" in sys.argv:
    target = xlsx_dir("data/demo")
elif args:
    target = args[0]
else:
    target = xlsx_dir("data/real")

files = sorted(glob.glob(os.path.join(target, "*.xlsx")))
if not files:
    print(f"[エラー] {target} に xlsx がありません。")
    print("  11台帳の xlsx を data/real/xlsx/ に置いてから実行してください(確認は --demo)。")
    sys.exit(1)
print(f"読み込み元: {target}  ({len(files)}ファイル)\n")

for fn in files:
    print(f"■ {os.path.basename(fn)}")
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    print(f"  シート数: {len(wb.sheetnames)}  シート名: {wb.sheetnames}")
    print(f"  アクティブシート(既定で読む対象): {wb.active.title}")
    for name in wb.sheetnames:
        ws = wb[name]
        # 先頭10行だけ見て、非空セルの多い行(見出し候補)を探す
        best_row_idx = None
        best_count = 0
        preview_rows = 0
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=10), start=1):
            preview_rows += 1
            non_empty = sum(1 for c in row if c is not None and str(c).strip() != "")
            if non_empty > best_count:
                best_count = non_empty
                best_row_idx = i
        print(f"    - シート「{name}」: 最大行={ws.max_row} 最大列={ws.max_column} "
              f"/ 先頭10行中もっとも列が埋まっている行= {best_row_idx}行目(非空{best_count}列)")
    wb.close()
    print()

# ── 範囲情報を無視して実データを数え直す(reset_dimensions) ──
print("== reset_dimensions で数え直し(壊れた範囲情報を無視して実際のセルを走査) ==")
print()
for fn in files:
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb.active
    ws.reset_dimensions()
    n_rows = 0
    n_cols_header = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None or all(c is None for c in row):
            continue
        if i == 1:
            n_cols_header = sum(1 for c in row if c is not None and str(c).strip() != "")
        n_rows += 1
    wb.close()
    print(f"  {os.path.basename(fn)}: 実データ行数(見出し込み)={n_rows:,}  見出し列数={n_cols_header}")

print()
# ── A1セルの「中身の形」だけを見る(個人情報は表示しない) ──
print("== A1セルの形チェック(内容は表示しません。文字数と区切り文字の有無のみ) ==")
print()
for fn in files:
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb.active
    a1 = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        a1 = row[0] if row else None
        break
    wb.close()
    if a1 is None:
        print(f"  {os.path.basename(fn)}: A1は空です")
        continue
    txt = str(a1)
    print(f"  {os.path.basename(fn)}: 型={type(a1).__name__} 文字数={len(txt)} "
          f"カンマ含む={',' in txt} タブ含む={chr(9) in txt} 改行含む={chr(10) in txt}")

print()
print("見るポイント:")
print("- 実データが入っていそうな行数・列数のシートが「アクティブシート」と一致しているか")
print("- 一致していない場合、そのシート名を教えてください(analyze/importの読み込み先を修正します)")
