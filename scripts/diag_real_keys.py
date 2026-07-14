# -*- coding: utf-8 -*-
"""顧客・商品の「本当の主キー」を確定する診断(個人情報は表示しない)。

宝飾ナビは店舗ごとに顧客ID/商品キーを採番している疑いがある。
その場合、主キーは「店舗コード + ID」の組。ここを確定しないと、
別店舗の別人を同じIDで上書きしてしまう。

表示するのは【件数・店舗コードの値・列見出し名】だけ。顧客名や電話は出さない。

使い方:
  python3 scripts/diag_real_keys.py            # data/real を見る
  python3 scripts/diag_real_keys.py --demo     # data/demo で確認
"""
import glob
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("[エラー] openpyxl が必要です。 pip3 install openpyxl")
    sys.exit(1)

target = "data/real"
if "--demo" in sys.argv:
    target = "data/demo"


def find(key):
    for p in sorted(glob.glob(os.path.join(target, "*.xlsx"))):
        if key in os.path.basename(p):
            return p
    return None


def open_ws(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    ws.reset_dimensions()
    return wb, ws


def col_index(header, *cands):
    for c in cands:
        for i, name in enumerate(header):
            if c in name:
                return i
    return None


def s(v):
    return "" if v is None else str(v).strip()


def analyze_key(path, label, id_cands):
    wb, ws = open_ws(path)
    it = ws.iter_rows(values_only=True)
    header = [s(h) for h in next(it)]
    si = col_index(header, "店舗コード", "店舗")
    ii = col_index(header, *id_cands)
    print(f"■ {label}: {os.path.basename(path)}")
    print(f"   店舗コード列= {header[si] if si is not None else '(なし)'} / "
          f"ID列= {header[ii] if ii is not None else '(なし)'}")
    rows = 0
    ids = set()
    pairs = set()
    per_store = defaultdict(int)
    for r in it:
        if r is None or all(c is None for c in r):
            continue
        rows += 1
        sid = s(r[si]) if si is not None and si < len(r) else ""
        cid = s(r[ii]) if ii is not None and ii < len(r) else ""
        ids.add(cid)
        pairs.add((sid, cid))
        per_store[sid] += 1
    wb.close()
    print(f"   総行数= {rows:,}")
    print(f"   IDの種類(単独)= {len(ids):,}")
    print(f"   (店舗コード+ID)の種類= {len(pairs):,}")
    if len(pairs) == rows:
        print("   → ★ 主キーは【店舗コード+ID】(この組で1行ずつ=完全にユニーク)")
    elif len(ids) == rows:
        print("   → 主キーはID単独でユニーク(店舗を跨いだ重複なし)")
    else:
        print("   → どちらでも完全一致せず。重複行あり(要精査)")
    print(f"   店舗コード別 行数:")
    for sid, n in sorted(per_store.items()):
        print(f"       店舗『{sid or '(空)'}』: {n:,}行")
    print()


def dump_headers(path, label):
    wb, ws = open_ws(path)
    it = ws.iter_rows(values_only=True)
    header = [s(h) for h in next(it)]
    wb.close()
    print(f"── {label} の列見出し({len(header)}列) ──")
    print("   " + " | ".join(f"{i}:{h}" for i, h in enumerate(header)))
    print()


def main():
    u = find("user")
    it_ = find("item")
    hb = find("hanbai")
    if u:
        analyze_key(u, "顧客台帳", ["顧客id"])
    if it_:
        analyze_key(it_, "商品台帳", ["商品キー", "商品コード", "商品id"])
    if hb:
        analyze_key(hb, "販売台帳", ["顧客id"])

    print("=== 主要台帳の列見出し(列名の確認用・個人情報ではありません) ===")
    print()
    for key, label in [("user", "顧客台帳"), ("item", "商品台帳"),
                       ("hanbai", "販売台帳"), ("shohosen", "処方箋台帳"),
                       ("point", "ポイント台帳"), ("pointhistory", "ポイント履歴")]:
        p = find(key)
        if p:
            dump_headers(p, label)


if __name__ == "__main__":
    main()
